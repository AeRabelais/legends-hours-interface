import pandas as pd
from openpyxl.styles import PatternFill
from openpyxl import Workbook
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
from reportlab.lib.enums import TA_LEFT, TA_CENTER
from reportlab.lib.units import inch
from legends_hours.settings import COLUMN_HEADERS


def create_excel_with_flags(data_frame: pd.DataFrame, file_path: str):
    # Create a new workbook
    workbook = Workbook()
    sheet = workbook.active

    # Write the column headers to the sheet
    column_headers = COLUMN_HEADERS
    for col_idx, header in enumerate(column_headers, start=1):
        sheet.cell(row=1, column=col_idx).value = header

    # Write the data rows to the sheet
    for row_idx, row in data_frame.iterrows():
        for col_idx, column in enumerate(column_headers, start=1):
            sheet.cell(row=row_idx + 2, column=col_idx).value = row[column]

            # Apply flag colors based on the 'flag' column
            if column == 'flag':
                flag_color = str(row['flag'])
                fill = PatternFill(patternType='solid',fgColor='FFFFFF')  # Default: white
                if flag_color == '1':
                    fill = PatternFill(patternType='solid',fgColor='FFFF00')  # Yellow
                elif flag_color == '2':
                    fill = PatternFill(patternType='solid', fgColor='FF0000')  # Red
                sheet.cell(row=row_idx + 2, column=col_idx).fill = fill

    # Save the workbook to the specified output file
    workbook.save(file_path)


def create_pdf_with_comments(data_frame: pd.DataFrame, file_path: str):

    start_date = str(data_frame['startDate'][0])

    # Create a list to store the content of each block
    content = []

    # Define the paragraph styles
    title_style = ParagraphStyle(
        name='Title',
        fontName='Helvetica-Bold',
        fontSize=18,
        leading=25,
        alignment=TA_CENTER
    )

    # Define the paragraph style
    paragraph_style = ParagraphStyle(
        name='Normal',
        fontName='Helvetica',
        fontSize=12,
        leading=15,
        leftIndent=10,
        alignment=TA_LEFT
    )

    # Add the title to the content list
    title = f"<u>Employee Overtime Comments for Week of {start_date}</u>"
    content.append(Paragraph(title, title_style))
    content.append(Paragraph("<br/><br/>", paragraph_style))  # Add some spacing

    # Iterate over the filtered DataFrame rows and add content to the list
    for _, row in data_frame.iterrows():
        employee_name = row['employee']
        total_hours = row['hours']
        comments = row['comment']
        num_over = row['num_overtime']

        block_content = f"<b>Employee:</b> {employee_name}, <b>Total Hours:</b> {total_hours},<b>Hours Overtime:</b> {num_over}"
        content.append(Paragraph(block_content, paragraph_style))
        comment_content = f"<b>Comments:</b> {comments}"
        content.append(Paragraph(comment_content, paragraph_style))
        content.append(Spacer(1, inch * 0.25))  # Add spacing between each employee block

    # Create the PDF document
    doc = SimpleDocTemplate(file_path, pagesize=letter)
    doc.build(content)
